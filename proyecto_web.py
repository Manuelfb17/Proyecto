import streamlit as st
import pandas as pd
from datetime import datetime
import holidays
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# ==============================
# Configuración inicial de sesión
# ==============================
if "registro_horas" not in st.session_state:
    st.session_state["registro_horas"] = {}  # Guarda todas las horas ingresadas

# ==============================
# ICONO Y NOMBRE PARA IOS (PWA)
# ==============================
st.markdown("""
<meta name="apple-mobile-web-app-title" content="Horas Extra Marco">
<link rel="apple-touch-icon" sizes="180x180" href="https://i.postimg.cc/7PjfgKkz/marco-peruana.png">
<meta name="apple-mobile-web-app-capable" content="yes">
""", unsafe_allow_html=True)

# ----------------------
# CONFIGURACIÓN DE LA PÁGINA
# ----------------------
st.set_page_config(
    page_title="Registro de Horas Extra",
    page_icon="⏰",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==============================
# ESTILOS: fondo dinámico y contenedor con leve blur
# ==============================
st.markdown(
    """
    <style>
    /* Fondo dinámico */
    .stApp {
        background: url('https://www.marco.com.pe/wp-content/uploads/2021/01/marco-7.jpg');
        background-size: cover;
        background-position: center;
        background-attachment: fixed;
    }

    /* Contenedor principal con blur leve al fondo */
    .contenido {
        margin-top: 20px;
        padding: 20px;
        border-radius: 10px;
        backdrop-filter: blur(8px); /* desenfoque solo del fondo */
        background-color: rgba(255,255,255,0.2); /* semi-transparente para ver fondo */
    }

    .campo-datos {
        margin-bottom: 20px;
    }
    </style>
    """, unsafe_allow_html=True
)

# ==============================
# CONTENIDO DE LA APP
# ==============================
with st.container():
    st.markdown('<div class="contenido"></div>', unsafe_allow_html=True)

    # BLOQUE DE DATOS GENERALES
    st.subheader("REGISTRO DE HORAS EXTRA")
    nombre_empleado = st.text_input("Ingrese su nombre", value="")
    sueldo_mensual = st.number_input(
        "Ingrese su sueldo mensual (S/):",
        min_value=0,
        step=100,
        format="%d",
        value=None
    )
    fecha_seleccionada = st.date_input("Seleccione la fecha (día, mes y año)", value=None)

    # BLOQUE HORAS EXTRA
    if fecha_seleccionada:
        fecha_str = fecha_seleccionada.strftime("%Y-%m-%d")
        peru_feriados = holidays.Peru(years=fecha_seleccionada.year)
        feriados = [fecha.strftime("%Y-%m-%d") for fecha in peru_feriados.keys()]

        st.markdown("<br><br>", unsafe_allow_html=True)
        st.subheader(f"Ingrese las horas extra para {fecha_str}")
        horas_extra = st.number_input(
            f"Horas extra del día seleccionado:",
            min_value=0,
            step=1,
            format="%d",
            value=st.session_state["registro_horas"].get(fecha_str, None)
        )
        st.session_state["registro_horas"][fecha_str] = horas_extra

    # BOTÓN CALCULAR Y TABLA
    if st.button("Calcular Horas Extra"):
        if nombre_empleado and sueldo_mensual:
            valor_hora = round(sueldo_mensual / (8 * 5 * 4.33), 2)
            registros = []

            for fecha_str, horas in st.session_state["registro_horas"].items():
                if horas:
                    fecha = datetime.strptime(fecha_str, "%Y-%m-%d")
                    dia_semana = fecha.weekday()  # 0=lunes, 6=domingo
                    es_domingo_o_feriado = (dia_semana == 5 or dia_semana == 6) or (fecha_str in feriados)

                    if es_domingo_o_feriado:
                        pago = round(horas * valor_hora * 2, 2)
                    else:
                        if horas <= 2:
                            pago = round(horas * valor_hora * 0.25, 2)
                        else:
                            pago = round(2*valor_hora*0.25 + (horas-2)*valor_hora*0.35, 2)

                    registros.append({
                        "Empleado": nombre_empleado,
                        "Fecha": fecha_str,
                        "Horas Extra": horas,
                        "Pago Extra (S/)": pago
                    })

            if registros:
                df = pd.DataFrame(registros)
                st.subheader("📊 Reporte de Horas Extra")
                st.dataframe(df)
                st.write("💰 **Total de horas extra (S/):**", df["Pago Extra (S/)"].sum())

                # Guardar Excel con columnas independientes y formateadas
                wb = Workbook()
                ws = wb.active
                ws.title = "Horas Extra"

                # Encabezados en negrita
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=col_num, value=column_title)
                    cell.font = Font(bold=True)

                # Datos
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

                # Ajuste ancho columnas
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[column].width = max_length + 2

                wb.save("HorasExtra_Mes_Reporte.xlsx")
                st.success("Reporte guardado como 'HorasExtra_Mes_Reporte.xlsx'")
            else:
                st.info("No se ingresaron horas extra.")
        else:
            st.warning("⚠️ Complete todos los campos.")
